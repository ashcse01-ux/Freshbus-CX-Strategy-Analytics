"""
sync_google_sheet.py
────────────────────
Fetches the Manual Metrics Google Sheet and merges the last 15 days of data
into the local 'Inbound Dashboard_Manual Metrics Data.xlsx' (sheet: Inbound Metrics).

Google Sheet: https://docs.google.com/spreadsheets/d/1ewwDxoCutZq_CKo9cJA8B9wHMTlOQPbd/edit?gid=741407078

Called automatically when the user clicks "Sync" on the dashboard.
- Updates existing date columns (so late-filled data gets refreshed)
- Appends brand new date columns
- Only processes the last 15 days
"""

import os
import sys
import io
import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ── CONFIG ────────────────────────────────────────────────────────────────────
GOOGLE_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/1ewwDxoCutZq_CKo9cJA8B9wHMTlOQPbd"
    "/gviz/tq?tqx=out:csv&gid=741407078"
)

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "additional_metrics_dump",
    "Inbound Dashboard_Manual Metrics Data.xlsx",
)
SHEET_NAME = "Inbound Metrics"

# Google Sheet row indices (0-based in the CSV)
GS_DAY_ROW  = 0   # Day names (Monday, Tuesday, …)
GS_DATE_ROW = 1   # Actual dates (11/1/2023, …)

# Mapping: Google Sheet metric name → local Excel row index (0-based, so +1 for openpyxl)
# Covers ALL metrics present in the Google Sheet including the 8 manual call metrics.
METRIC_MAP = {
    "Gross Seats":                        2,
    "Gross Tickets":                      3,
    "Intr/Journey ( Overall )":           4,
    "Present Agent HC":                   7,
    "Intr/Journey % ( Inbound + WH)":     8,
    "Intr/Journey %  ( Travel update )":  9,
    "No. of Service Delay":              10,
    "Delay Pax Impacted":                11,
    "No. of Service Cancel":             12,
    "Service Cancel Pax Impacted":       13,
    "No. of Service Breakdown":          14,
    "Break Down Pax Impacted":           15,
    "Impacted %":                        16,
    "Total Pax Impacted":                17,
    "Cancellations Impact %":            18,
    # ── 8 manual call metrics ──────────────────────────────────────────────
    "Call Drop Not Done":                20,
    "Blank Call Not Done":               21,
    "Overall Call Not Done":             22,
    "Call Not Done %":                   23,
    "Agent Disconnected":                25,
    "Agent Disconnected %":              26,
    "Call Not Disposed":                 28,
    "Call Not Disposed %":               29,
}


def _parse_date(val):
    """Parse a date string from the Google Sheet into a datetime object."""
    if pd.isna(val) or not str(val).strip():
        return None
    raw = str(val).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    try:
        return pd.to_datetime(raw).to_pydatetime()
    except Exception:
        return None


def _parse_cell_value(val):
    """Convert a Google Sheet cell value to a clean Python value."""
    if pd.isna(val) or str(val).strip() == "":
        return None
    s = str(val).strip()
    # Percentage string → decimal float
    if s.endswith("%"):
        try:
            return float(s.replace("%", "").replace(",", "")) / 100.0
        except ValueError:
            return s
    # Plain number
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return s


def sync():
    """Main entry point — fetch Google Sheet and merge into local Excel."""

    # ── 1. Fetch Google Sheet ────────────────────────────────────────────────
    print("Fetching Google Sheet data...")
    try:
        resp = requests.get(GOOGLE_SHEET_URL, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"  Failed to fetch Google Sheet: {e}")
        return 0

    gs_df = pd.read_csv(io.StringIO(resp.text), header=None)
    print(f"  Google Sheet shape: {gs_df.shape}")

    # ── 2. Build dict: date_str → column index in Google Sheet ───────────────
    gs_date_cols = {}
    for j in range(1, gs_df.shape[1]):
        dt = _parse_date(gs_df.iloc[GS_DATE_ROW, j])
        if dt:
            gs_date_cols[dt.strftime("%Y-%m-%d")] = j

    print(f"  Google Sheet dates: {len(gs_date_cols)}  "
          f"({min(gs_date_cols)} → {max(gs_date_cols)})")

    # ── 3. Build dict: Google Sheet metric name → row index in gs_df ─────────
    gs_metric_rows = {}
    for i in range(gs_df.shape[0]):
        name = gs_df.iloc[i, 0]
        if pd.notna(name) and str(name).strip():
            gs_metric_rows[str(name).strip()] = i

    # ── 4. Calculate cutoff date (last 15 days from current date) ────────────
    today = datetime.now()
    cutoff_str = (today - timedelta(days=15)).strftime("%Y-%m-%d")
    today_str = today.strftime("%Y-%m-%d")
    print(f"  Only syncing dates from {cutoff_str} to {today_str} (last 15 days).")

    # ── 5. Load local Excel and find existing dates ──────────────────────────
    if not os.path.exists(EXCEL_PATH):
        print(f"  Local Excel not found: {EXCEL_PATH}")
        return 0

    local_df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, header=None)
    print(f"  Local Excel shape: {local_df.shape}")

    # ── 6. Open workbook with openpyxl ───────────────────────────────────────
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # Find the last active column (row 1 or row 2 is non-empty)
    last_active_col = 2
    for j in range(ws.max_column, 1, -1):
        v1 = ws.cell(row=1, column=j).value
        v2 = ws.cell(row=2, column=j).value
        if (v1 is not None and str(v1).strip() != "") or (v2 is not None and str(v2).strip() != ""):
            last_active_col = j
            break

    # Remove any empty trailing columns beyond the last active one
    if ws.max_column > last_active_col:
        print(f"  Cleaning up {ws.max_column - last_active_col} empty trailing columns...")
        ws.delete_cols(last_active_col + 1, ws.max_column - last_active_col)

    # Build local date → column mapping (scan only last 50 cols to avoid misaligned template dates)
    local_date_cols = {}
    start_col = max(2, ws.max_column - 50)
    for j in range(ws.max_column, start_col - 1, -1):
        dt_val = ws.cell(row=2, column=j).value
        if pd.notna(dt_val):
            try:
                dt = pd.to_datetime(dt_val)
                date_str = dt.strftime("%Y-%m-%d")
                if date_str not in local_date_cols:
                    local_date_cols[date_str] = j
            except Exception:
                pass

    print(f"  Local date columns (last 50 cols): {len(local_date_cols)}")

    # ── 7. Write data for each date in the last 15 days ──────────────────────
    added = 0
    updated = 0
    next_col = ws.max_column + 1

    for date_str, gs_col in sorted(gs_date_cols.items()):
        # Skip dates older than 15 days or future dates
        if date_str < cutoff_str or date_str > today_str:
            continue

        dt = datetime.strptime(date_str, "%Y-%m-%d")

        if date_str in local_date_cols:
            col_idx = local_date_cols[date_str]
            updated += 1
        else:
            col_idx = next_col
            next_col += 1
            added += 1

        # Row 1: Day name
        day_name = gs_df.iloc[GS_DAY_ROW, gs_col]
        ws.cell(row=1, column=col_idx, value=str(day_name) if pd.notna(day_name) else "")

        # Row 2: Date (as datetime for proper Excel formatting)
        ws.cell(row=2, column=col_idx, value=dt)

        # All metric rows
        for metric_name, local_row_idx in METRIC_MAP.items():
            gs_row_idx = gs_metric_rows.get(metric_name)
            if gs_row_idx is not None:
                raw_val = gs_df.iloc[gs_row_idx, gs_col]
                cell_val = _parse_cell_value(raw_val)
                # Write value (None clears the cell if previously set)
                ws.cell(row=local_row_idx + 1, column=col_idx, value=cell_val)

    # ── 8. Save ──────────────────────────────────────────────────────────────
    wb.save(EXCEL_PATH)
    print(f"  Done: {updated} dates updated, {added} new dates appended.")
    return added + updated


if __name__ == "__main__":
    sync()
