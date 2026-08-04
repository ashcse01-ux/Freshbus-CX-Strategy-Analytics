import sqlite3
import os
from openpyxl import load_workbook

def iter_rows(ws, cols_to_extract):
    """
    Yields dicts mapping target column name to cell value.
    cols_to_extract is a dict { "Excel Header Name": "db_column_name" }
    """
    header = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            for col_idx, val in enumerate(row):
                if val and str(val).strip() in cols_to_extract:
                    header[col_idx] = cols_to_extract[str(val).strip()]
            continue
        
        # Build dict for row
        row_data = {}
        has_data = False
        for col_idx, db_col in header.items():
            if col_idx < len(row):
                val = row[col_idx]
                row_data[db_col] = val
                if val is not None and str(val).strip() != '':
                    has_data = True
            else:
                row_data[db_col] = None
                
        if has_data:
            yield row_data

def run_sync():
    base_dir = os.path.dirname(os.path.dirname(__file__))
    excel_path = os.path.join(base_dir, "Redbus Analytics Dashboard Preloaded Data Dump", "Redbus Dashboard for Automation - 1st June to 14th July.xlsx")
    db_path = os.path.join(base_dir, "backend", "metrics_redbus.db")

    print(f"Loading workbook (read_only)... {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)

    print(f"Connecting to DB: {db_path}")
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Setup Tables
    cursor.execute("DROP TABLE IF EXISTS call_sheet")
    cursor.execute("""
    CREATE TABLE call_sheet (
        doj TEXT, pnr TEXT, agent_name TEXT, month TEXT, call_status TEXT, tl_name TEXT
    )""")
    
    cursor.execute("DROP TABLE IF EXISTS rating_dump")
    cursor.execute("""
    CREATE TABLE rating_dump (
        pnr TEXT, route TEXT, doj TEXT, rating_date TEXT, rating REAL, call_status TEXT
    )""")
    
    cursor.execute("DROP TABLE IF EXISTS travel_data")
    cursor.execute("""
    CREATE TABLE travel_data (
        pnr TEXT, doj TEXT, route TEXT
    )""")

    print("Parsing call sheet...")
    call_cols = {'DOJ Correct': 'doj', 'PNR': 'pnr', 'Agnet Name': 'agent_name', 'Month': 'month', 'Call Status': 'call_status', 'Tl Names': 'tl_name'}
    ws = wb['call sheet']
    for row in iter_rows(ws, call_cols):
        cursor.execute("INSERT INTO call_sheet (doj, pnr, agent_name, month, call_status, tl_name) VALUES (?, ?, ?, ?, ?, ?)",
            (str(row.get('doj','')), str(row.get('pnr','')), str(row.get('agent_name','')), str(row.get('month','')), str(row.get('call_status','')), str(row.get('tl_name','')))
        )

    print("Parsing rating dump...")
    rating_cols = {'PNR': 'pnr', 'Route': 'route', 'DateOfJourney': 'doj', 'DateOfRating': 'rating_date', 'Rating': 'rating', 'Call Status': 'call_status'}
    ws = wb['rating Dump']
    for row in iter_rows(ws, rating_cols):
        rating_val = None
        try:
            if row.get('rating') is not None:
                rating_val = float(row.get('rating'))
        except ValueError:
            pass
        cursor.execute("INSERT INTO rating_dump (pnr, route, doj, rating_date, rating, call_status) VALUES (?, ?, ?, ?, ?, ?)",
            (str(row.get('pnr','')), str(row.get('route','')), str(row.get('doj','')), str(row.get('rating_date','')), rating_val, str(row.get('call_status','')))
        )

    print("Parsing travel data...")
    travel_cols = {'Ticket No': 'pnr', 'Journey Date': 'doj', 'Route': 'route'}
    ws = wb['Travel Data']
    for row in iter_rows(ws, travel_cols):
        cursor.execute("INSERT INTO travel_data (pnr, doj, route) VALUES (?, ?, ?)",
            (str(row.get('pnr','')), str(row.get('doj','')), str(row.get('route','')))
        )

    print("Building redbus_master table...")
    cursor.execute("DROP TABLE IF EXISTS redbus_master")
    
    create_master_sql = """
    CREATE TABLE redbus_master AS
    SELECT 
        UPPER(TRIM(t.pnr)) AS pnr,
        substr(t.doj, 1, 10) AS date,
        t.route AS route_name,
        
        CASE WHEN c.pnr IS NOT NULL THEN 1 ELSE 0 END as is_assigned,
        TRIM(c.tl_name) AS tl_name,
        TRIM(c.agent_name) AS agent_name,
        TRIM(LOWER(c.call_status)) AS call_status,
        
        CASE WHEN r.pnr IS NOT NULL THEN 1 ELSE 0 END as is_responded,
        r.rating,
        
        CASE 
            WHEN r.pnr IS NOT NULL AND (c.pnr IS NULL OR TRIM(LOWER(c.call_status)) = 'not connected' OR c.call_status IS NULL OR c.call_status = '') THEN 'Organic'
            WHEN r.pnr IS NOT NULL AND TRIM(LOWER(c.call_status)) = 'connected' THEN 'Inorganic'
            ELSE 'None'
        END AS review_type
        
    FROM travel_data t
    LEFT JOIN call_sheet c ON UPPER(TRIM(t.pnr)) = UPPER(TRIM(c.pnr))
    LEFT JOIN rating_dump r ON UPPER(TRIM(t.pnr)) = UPPER(TRIM(r.pnr))
    """
    cursor.execute(create_master_sql)

    insert_missing_sql = """
    INSERT INTO redbus_master (pnr, date, route_name, is_assigned, tl_name, agent_name, call_status, is_responded, rating, review_type)
    SELECT 
        UPPER(TRIM(r.pnr)) AS pnr,
        substr(r.doj, 1, 10) AS date,
        r.route AS route_name,
        CASE WHEN c.pnr IS NOT NULL THEN 1 ELSE 0 END as is_assigned,
        TRIM(c.tl_name) AS tl_name,
        TRIM(c.agent_name) AS agent_name,
        TRIM(LOWER(c.call_status)) AS call_status,
        1 as is_responded,
        r.rating,
        CASE 
            WHEN c.pnr IS NULL OR TRIM(LOWER(c.call_status)) = 'not connected' OR c.call_status IS NULL OR c.call_status = '' THEN 'Organic'
            WHEN TRIM(LOWER(c.call_status)) = 'connected' THEN 'Inorganic'
            ELSE 'None'
        END AS review_type
    FROM rating_dump r
    LEFT JOIN call_sheet c ON UPPER(TRIM(r.pnr)) = UPPER(TRIM(c.pnr))
    WHERE UPPER(TRIM(r.pnr)) NOT IN (SELECT pnr FROM redbus_master)
    """
    cursor.execute(insert_missing_sql)

    conn.commit()
    conn.close()
    wb.close()
    print("Database sync completed successfully via openpyxl!")

if __name__ == "__main__":
    run_sync()
